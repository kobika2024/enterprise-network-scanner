"""
Excel Report Generator — 5 RTL worksheets using openpyxl.

Sheets:
  1. מלאי נכסים   — Asset Inventory (main table)
  2. פורטים        — Open Ports detail
  3. סטטיסטיקות   — Summary statistics
  4. Docker         — Docker container details
  5. SNMP           — SNMP-discovered device details
"""
from io import BytesIO
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side,
        numbers as xl_numbers,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# -----------------------------------------------------------------------
# Style helpers
# -----------------------------------------------------------------------

def _make_header_fill(color='1F4E79'):
    return PatternFill('solid', fgColor=color)


def _make_font(bold=False, color='000000', size=11):
    return Font(bold=bold, color=color, size=size, name='Calibri')


def _make_border():
    thin = Side(style='thin', color='BFBFBF')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _risk_fill(risk: str) -> PatternFill:
    color_map = {
        'CRITICAL': 'FF0000',
        'HIGH':     'FF6600',
        'MEDIUM':   'FFCC00',
        'LOW':      '92D050',
        'INFO':     'DEEBF7',
    }
    color = color_map.get(risk, 'FFFFFF')
    return PatternFill('solid', fgColor=color)


def _asset_type_fill(asset_type: str) -> PatternFill:
    color_map = {
        'Server':           'BDD7EE',
        'Domain Controller':'9BC2E6',
        'Database Server':  'B4C7E7',
        'Web Server':       'C6EFCE',
        'Workstation':      'E2EFDA',
        'VM':               'FFF2CC',
        'Network Device':   'FCE4D6',
        'Printer':          'F2DCDB',
        'Container Host':   'E2EFDA',
        'Hypervisor':       'D9D9D9',
    }
    c = color_map.get(asset_type, 'FFFFFF')
    return PatternFill('solid', fgColor=c)


def _style_header_row(ws, row: int, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _make_header_fill()
        cell.font = _make_font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True, reading_order=2)
        cell.border = _make_border()


def _style_data_row(ws, row: int, num_cols: int, fill=None):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        if fill:
            cell.fill = fill
        cell.font = _make_font(size=10)
        cell.alignment = Alignment(horizontal='right', vertical='center',
                                   reading_order=2)
        cell.border = _make_border()


def _set_rtl(ws):
    ws.sheet_view.rightToLeft = True


def _auto_width(ws, min_width=10, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, min_width), max_width)


# -----------------------------------------------------------------------
# Sheet 1: Asset Inventory
# -----------------------------------------------------------------------

ASSET_HEADERS = [
    'IP', 'MAC', 'ספק', 'שם מחשב', 'FQDN',
    'משפחת OS', 'גרסת OS', 'Build', 'שיטת זיהוי',
    'סוג נכס', 'פורטים פתוחים', 'שיטות גילוי',
    'Docker', 'הראשון נראה', 'אחרון נראה',
]


def _sheet_assets(wb, assets: list):
    ws = wb.create_sheet('מלאי נכסים')
    _set_rtl(ws)
    ws.freeze_panes = 'A2'

    # Header
    for col, h in enumerate(ASSET_HEADERS, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(ASSET_HEADERS))
    ws.row_dimensions[1].height = 30

    for row_idx, a in enumerate(assets, 2):
        open_ports_str = ', '.join(
            str(p.get('port', '')) for p in (a.get('open_ports') or [])
        )
        containers_count = len(a.get('docker_containers') or [])
        docker_str = f'כן ({containers_count})' if a.get('is_docker_host') else ''

        values = [
            a.get('ip', ''),
            a.get('mac', ''),
            a.get('mac_vendor', ''),
            a.get('hostname', ''),
            a.get('fqdn', ''),
            a.get('os_family', ''),
            a.get('os_version', ''),
            a.get('os_build', ''),
            a.get('os_method', ''),
            a.get('asset_type', ''),
            open_ports_str,
            ', '.join(a.get('discovery_methods') or []),
            docker_str,
            a.get('first_seen', '')[:19] if a.get('first_seen') else '',
            a.get('last_seen', '')[:19] if a.get('last_seen') else '',
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=val)

        fill = _asset_type_fill(a.get('asset_type', ''))
        # Highlight critical/high risk rows
        if a.get('critical_ports', 0) > 0:
            fill = PatternFill('solid', fgColor='FFCCCC')
        elif a.get('high_ports', 0) > 0:
            fill = PatternFill('solid', fgColor='FFE6CC')

        _style_data_row(ws, row_idx, len(ASSET_HEADERS), fill=fill)

    _auto_width(ws)


# -----------------------------------------------------------------------
# Sheet 2: Ports
# -----------------------------------------------------------------------

PORT_HEADERS = ['IP', 'פורט', 'שירות', 'סיכון', 'באנר']


def _sheet_ports(wb, assets: list):
    ws = wb.create_sheet('פורטים')
    _set_rtl(ws)
    ws.freeze_panes = 'A2'

    for col, h in enumerate(PORT_HEADERS, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(PORT_HEADERS))

    row_idx = 2
    for a in assets:
        ip = a.get('ip', '')
        for port in (a.get('open_ports') or []):
            risk = port.get('risk', 'INFO')
            values = [
                ip,
                port.get('port', ''),
                port.get('service', ''),
                risk,
                port.get('banner', '')[:120],
            ]
            for col, val in enumerate(values, 1):
                ws.cell(row=row_idx, column=col, value=val)
            fill = _risk_fill(risk)
            _style_data_row(ws, row_idx, len(PORT_HEADERS), fill=fill)
            row_idx += 1

    _auto_width(ws)


# -----------------------------------------------------------------------
# Sheet 3: Statistics
# -----------------------------------------------------------------------

def _sheet_stats(wb, assets: list, meta: dict):
    ws = wb.create_sheet('סטטיסטיקות')
    _set_rtl(ws)

    # --- Meta section ---
    ws['A1'] = 'פרמטר'
    ws['B1'] = 'ערך'
    _style_header_row(ws, 1, 2)

    meta_rows = [
        ('subnet/target', meta.get('target', '')),
        ('תאריך סריקה', meta.get('scan_date', datetime.now().strftime('%Y-%m-%d %H:%M'))),
        ('משך סריקה', meta.get('duration', '')),
        ('סה"כ נכסים', len(assets)),
        ('נכסים חיים', sum(1 for a in assets if a.get('is_alive'))),
        ('פורטים קריטיים', sum(a.get('critical_ports', 0) for a in assets)),
    ]
    for i, (k, v) in enumerate(meta_rows, 2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        _style_data_row(ws, i, 2)

    # --- OS count table ---
    os_counts = {}
    for a in assets:
        fam = a.get('os_family') or 'Unknown'
        os_counts[fam] = os_counts.get(fam, 0) + 1

    start_row = len(meta_rows) + 3
    ws.cell(row=start_row, column=1, value='משפחת OS')
    ws.cell(row=start_row, column=2, value='כמות')
    _style_header_row(ws, start_row, 2)

    for i, (os, count) in enumerate(sorted(os_counts.items(), key=lambda x: -x[1]), start_row + 1):
        ws.cell(row=i, column=1, value=os)
        ws.cell(row=i, column=2, value=count)
        _style_data_row(ws, i, 2)

    # --- Asset type count table ---
    type_start = start_row + len(os_counts) + 3
    type_counts = {}
    for a in assets:
        t = a.get('asset_type') or 'Unknown'
        type_counts[t] = type_counts.get(t, 0) + 1

    ws.cell(row=type_start, column=1, value='סוג נכס')
    ws.cell(row=type_start, column=2, value='כמות')
    _style_header_row(ws, type_start, 2)

    for i, (t, count) in enumerate(sorted(type_counts.items(), key=lambda x: -x[1]), type_start + 1):
        ws.cell(row=i, column=1, value=t)
        ws.cell(row=i, column=2, value=count)
        _style_data_row(ws, i, 2)

    # Bar chart for asset types
    try:
        chart = BarChart()
        chart.type = 'col'
        chart.title = 'סוגי נכסים'
        chart.y_axis.title = 'כמות'
        chart.x_axis.title = 'סוג'

        data_ref = Reference(ws,
                             min_col=2, max_col=2,
                             min_row=type_start, max_row=type_start + len(type_counts))
        cats_ref = Reference(ws,
                             min_col=1,
                             min_row=type_start + 1, max_row=type_start + len(type_counts))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 15
        chart.height = 10
        ws.add_chart(chart, f'D{start_row}')
    except Exception:
        pass

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15


# -----------------------------------------------------------------------
# Sheet 4: Docker Containers
# -----------------------------------------------------------------------

DOCKER_HEADERS = ['Host IP', 'Container ID', 'שם', 'Image', 'סטטוס', 'פורטים']


def _sheet_docker(wb, assets: list):
    ws = wb.create_sheet('Docker Containers')
    _set_rtl(ws)

    for col, h in enumerate(DOCKER_HEADERS, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(DOCKER_HEADERS))

    row_idx = 2
    for a in assets:
        if not a.get('is_docker_host'):
            continue
        ip = a.get('ip', '')
        for c in (a.get('docker_containers') or []):
            values = [
                ip,
                c.get('id', ''),
                c.get('name', ''),
                c.get('image', ''),
                c.get('status', ''),
                ', '.join(c.get('ports') or []),
            ]
            for col, val in enumerate(values, 1):
                ws.cell(row=row_idx, column=col, value=val)
            _style_data_row(ws, row_idx, len(DOCKER_HEADERS))
            row_idx += 1

    _auto_width(ws)


# -----------------------------------------------------------------------
# Sheet 5: SNMP Devices
# -----------------------------------------------------------------------

SNMP_HEADERS = ['IP', 'sysName', 'sysDescr', 'מיקום', 'ממשקים']


def _sheet_snmp(wb, assets: list):
    ws = wb.create_sheet('SNMP')
    _set_rtl(ws)

    for col, h in enumerate(SNMP_HEADERS, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(SNMP_HEADERS))

    row_idx = 2
    for a in assets:
        if not a.get('snmp_sysname'):
            continue
        interfaces_str = '; '.join(
            f"{i.get('descr', '')} {i.get('speed', '')} {i.get('mac', '')}"
            for i in (a.get('snmp_interfaces') or [])
        )[:200]
        values = [
            a.get('ip', ''),
            a.get('snmp_sysname', ''),
            (a.get('snmp_sysdesc') or '')[:150],
            a.get('snmp_location', ''),
            interfaces_str,
        ]
        for col, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=val)
        _style_data_row(ws, row_idx, len(SNMP_HEADERS))
        row_idx += 1

    _auto_width(ws)


# -----------------------------------------------------------------------
# Public API
# -----------------------------------------------------------------------

def generate_excel(assets: list, meta: dict = None) -> bytes:
    """
    Generate a full Excel report from asset list.

    Args:
        assets: List of Asset.to_dict() results
        meta:   Optional dict with {target, scan_date, duration}

    Returns:
        Raw .xlsx bytes
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError('openpyxl is not installed. Run: pip install openpyxl')

    if meta is None:
        meta = {}

    wb = openpyxl.Workbook()
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    _sheet_assets(wb, assets)
    _sheet_ports(wb, assets)
    _sheet_stats(wb, assets, meta)
    _sheet_docker(wb, assets)
    _sheet_snmp(wb, assets)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
